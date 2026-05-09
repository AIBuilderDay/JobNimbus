"""Convert uploaded files (PDF / CSV / Excel / image) into Anthropic
content blocks the agent can reason over.

PDFs and images go in as their native multimodal blocks (`document` /
`image`). CSV and Excel get parsed server-side into compact text, which
Claude reads alongside the user's message — sending raw bytes for
spreadsheets would waste tokens and Claude can't natively parse XLSX.

We log the size and shape of every parsed file so when extraction fails
the failure surfaces in the backend log next to the SSE error event.
"""
from __future__ import annotations

import base64
import csv
import io
from typing import Any

from logger import get_logger

log = get_logger(__name__)

# Content-type allow-list. Order matters — we check by suffix first, then mime.
SUPPORTED_PDF = {"application/pdf"}
SUPPORTED_IMAGE = {"image/jpeg", "image/png", "image/gif", "image/webp"}
SUPPORTED_CSV = {"text/csv", "application/csv"}
SUPPORTED_XLSX = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/octet-stream",  # browsers sometimes send xlsx as octet-stream
}

# Cap how much spreadsheet text we send — anything past this gets truncated
# with a note. PDFs are sent as documents so size is enforced by the API.
MAX_TABULAR_CHARS = 50_000


class FilePart:
    """A single uploaded file plus its derived Anthropic content blocks."""

    def __init__(
        self,
        filename: str,
        kind: str,
        blocks: list[dict[str, Any]],
        summary: str,
    ) -> None:
        self.filename = filename
        self.kind = kind  # "pdf" | "image" | "csv" | "xlsx"
        self.blocks = blocks
        self.summary = summary

    def to_dict(self) -> dict[str, Any]:
        return {"filename": self.filename, "kind": self.kind, "summary": self.summary}


def parse_upload(filename: str, content_type: str | None, data: bytes) -> FilePart:
    """Return the FilePart for an uploaded file. Raises ValueError on unknown type."""
    name = (filename or "").lower()
    ct = (content_type or "").lower()

    if ct in SUPPORTED_PDF or name.endswith(".pdf"):
        return _pdf_part(filename, data)
    if ct in SUPPORTED_IMAGE or any(name.endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".gif", ".webp")):
        return _image_part(filename, ct, data)
    if ct in SUPPORTED_CSV or name.endswith(".csv"):
        return _csv_part(filename, data)
    if ct in SUPPORTED_XLSX or name.endswith(".xlsx") or name.endswith(".xls"):
        return _xlsx_part(filename, data)

    log.warning("upload rejected filename=%s content_type=%s size=%d", filename, ct, len(data))
    raise ValueError(
        f"Unsupported file type: {filename!r} ({ct or 'no content-type'}). "
        "Accepted: PDF, CSV, XLSX, JPG, PNG, GIF, WEBP."
    )


def _pdf_part(filename: str, data: bytes) -> FilePart:
    log.info("upload pdf filename=%s size=%d", filename, len(data))
    b64 = base64.standard_b64encode(data).decode("ascii")
    blocks: list[dict[str, Any]] = [
        {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
            "title": filename,
        },
        {"type": "text", "text": f"(Attached PDF: {filename})"},
    ]
    return FilePart(filename, "pdf", blocks, summary=f"PDF · {len(data) // 1024} KB")


def _image_part(filename: str, content_type: str, data: bytes) -> FilePart:
    media_type = content_type if content_type in SUPPORTED_IMAGE else "image/jpeg"
    log.info("upload image filename=%s media_type=%s size=%d", filename, media_type, len(data))
    b64 = base64.standard_b64encode(data).decode("ascii")
    blocks: list[dict[str, Any]] = [
        {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
        {"type": "text", "text": f"(Attached image: {filename})"},
    ]
    return FilePart(filename, "image", blocks, summary=f"Image · {len(data) // 1024} KB")


def _csv_part(filename: str, data: bytes) -> FilePart:
    text = data.decode("utf-8", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    rendered = _render_table(filename, rows)
    log.info("upload csv filename=%s rows=%d truncated=%s", filename, len(rows), len(rendered) > MAX_TABULAR_CHARS)
    blocks = [{"type": "text", "text": rendered}]
    return FilePart(filename, "csv", blocks, summary=f"CSV · {len(rows)} rows")


def _xlsx_part(filename: str, data: bytes) -> FilePart:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        log.exception("openpyxl not available — cannot parse xlsx")
        raise ValueError("Server is missing openpyxl; cannot parse Excel files.") from e

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — openpyxl raises lots of types
        log.exception("xlsx parse failed filename=%s", filename)
        raise ValueError(f"Could not parse Excel file {filename!r}: {e}") from e

    chunks: list[str] = []
    total_rows = 0
    for sheet in wb.worksheets:
        rows: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(["" if v is None else v for v in row])
            total_rows += 1
        chunks.append(_render_table(f"{filename} · sheet={sheet.title}", rows))

    rendered = "\n\n".join(chunks)
    if len(rendered) > MAX_TABULAR_CHARS:
        rendered = rendered[:MAX_TABULAR_CHARS] + f"\n\n…(truncated; full file was {total_rows} rows)"
    log.info(
        "upload xlsx filename=%s sheets=%d rows=%d size=%d",
        filename, len(wb.worksheets), total_rows, len(data),
    )
    blocks = [{"type": "text", "text": rendered}]
    return FilePart(filename, "xlsx", blocks, summary=f"Excel · {total_rows} rows · {len(wb.worksheets)} sheet(s)")


def _render_table(label: str, rows: list[list[Any]]) -> str:
    """Render a 2D list as a fenced markdown-like block. Compact, deterministic.
    No fancy padding — Claude reads tabs just fine and we save tokens."""
    body = "\n".join("\t".join(str(c) for c in r) for r in rows)
    rendered = f"--- {label} ---\n{body}\n--- end {label} ---"
    if len(rendered) > MAX_TABULAR_CHARS:
        rendered = rendered[:MAX_TABULAR_CHARS] + f"\n…(truncated; total rows={len(rows)})"
    return rendered
